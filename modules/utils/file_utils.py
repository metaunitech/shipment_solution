from pathlib import Path
import pandas as pd
import xlwings as xw
from openpyxl import load_workbook
import PyPDF2
import fitz
import chardet
import re
from PIL import Image
import os, io

"""
File type transformation
"""


def transform_csv2xlsx(file_path: Path):
    """
    Transform CSV file to XLSX format.
    :param file_path: Path to the CSV file.
    """
    csv_data = pd.read_csv(file_path)  # 读取 CSV 文件
    xlsx_path = file_path.with_suffix('.xlsx')  # 生成相应的 XLSX 文件路径
    csv_data.to_excel(xlsx_path, index=False)  # 将数据保存为 XLSX 文件
    return xlsx_path


def transform_xls2xlsx(file_path: Path):
    """
    Transform XLS file to XLSX format.
    :param file_path: Path to the XLS file.
    """
    xls_data = pd.read_excel(file_path, header=None, engine='xlrd')  # 读取 XLS 文件
    xls_data = xls_data.dropna(axis=1, how='all')
    xlsx_path = file_path.with_suffix('.xlsx')  # 生成相应的 XLSX 文件路径
    xls_data.to_excel(xlsx_path, index=False)  # 将数据保存为 XLSX 文件
    return xlsx_path


def transform_et2xlsx(file_path: Path):
    """
    Transform ET file to XLSX format.
    :param file_path: Path to the ET file.
    :return: Path to the generated XLSX file.
    """
    # 使用 xlwings 打开 ET 文件
    app = xw.App(visible=False)
    wb = app.books.open(file_path)

    # 读取所有数据
    data = []
    for sheet in wb.sheets:
        data.extend(sheet.used_range.value)

    # 关闭 ET 文件
    wb.close()
    app.quit()

    # 生成 XLSX 文件路径
    xlsx_path = file_path.with_suffix('.xlsx')

    # 将数据转换为 DataFrame
    df = pd.DataFrame(data)

    # 保存为 xlsx 文件
    df.to_excel(xlsx_path, index=False)
    return xlsx_path


def extract_plaintext_from_xlsx(file_path: Path):
    texts = []

    # 加载xlsx文件
    workbook = load_workbook(file_path)

    # 选择第一个工作表（可以根据需要修改索引）
    sheet = workbook.active

    # 逐行提取文本
    for row in sheet.iter_rows(min_row=1, max_col=sheet.max_column, max_row=sheet.max_row, values_only=True):
        # 将每行的文本连接起来，使用制表符分隔
        if all(cell is None for cell in row):
            print("Empty row, skipped")
        row_text = ' '.join(str(cell) if cell is not None else ' ' for cell in row)

        # 将当前行的文本添加到列表中
        texts.append(row_text)

    # 关闭工作簿
    workbook.close()

    # 返回包含所有文本的列表
    return texts


"""
General
"""


def detect_encoding(text):
    result = chardet.detect(text)
    encoding = result['encoding']

    return encoding


"""
PDF handler
"""


def is_scanned_pdf(file_path):
    with open(file_path, 'rb') as pdf_file:
        pdf_reader = PyPDF2.PdfReader(pdf_file)

        for page_num in range(len(pdf_reader.pages)):
            page = pdf_reader.pages[page_num]
            page_text = page.extract_text()
            contain_unknown_encoding = len(re.findall(r'[\x80-\xFF]+', page_text)) >= 10
            if contain_unknown_encoding:
                return True
        return False


def extract_images_from_pdf(input_file_path, storage_path):
    # open file
    with fitz.Document(input_file_path) as my_pdf_file:
        # 遍历所有页面
        imgs = {}
        for page_number in range(1, len(my_pdf_file) + 1):
            # 查看独立页面
            page = my_pdf_file[page_number - 1]
            # # 查看当前页所有图片
            # images = page.get_images()
            # 遍历当前页面所有图片
            for image_number, image in enumerate(page.get_images(), start=1):
                # 访问图片xref
                xref_value = image[0]
                # image name
                image_name = image[-2]
                # 提取图片信息
                base_image = my_pdf_file.extract_image(xref_value)
                # 访问图片
                image_bytes = base_image["image"]
                # 获取图片扩展名
                ext = base_image["ext"]
                # 加载图片
                image = Image.open(io.BytesIO(image_bytes))

                imgs[f"page{str(page_number)}_{image_name}"] = image
    output = {}

    for image_name in imgs.keys():
        image_file_name = f"image_{image_name}.{ext}"
        im_path = os.path.join(storage_path, image_file_name)
        image = imgs[image_name]
        image.save(open(im_path, "wb"))
        output[image_name] = im_path
    return output


def extract_plaintext_from_pdf(input_file_path):
    with fitz.Document(input_file_path) as my_pdf_file:
        # 遍历所有页面
        texts = []
        text_list = [page.get_text() for page in my_pdf_file]
        all_text = ''
        for text in text_list:
            all_text += text
        for line in all_text.split('\n'):
            texts.append(line)
    return texts


def convert_pdf2img(input_file_path: Path, storage_path=None):
    if not storage_path:
        storage_path = Path(__file__).parent
    page_num = 1
    filename = input_file_path.stem
    os.makedirs(storage_path / filename, exist_ok=True)
    with fitz.Document(input_file_path) as pdf:
        img_outs = {}
        for page in pdf:
            zoom_x = 2
            zoom_y = 2
            mat = fitz.Matrix(zoom_x, zoom_y)
            pixmap = page.get_pixmap(matrix=mat, alpha=False)
            image_file = storage_path / f"{filename}/{page_num}.png"
            # if image_file.exists():
            #     print(f"第{page_num}保存图片完成")
            #     continue
            pixmap.pil_save(image_file)
            img_outs[page_num] = image_file
            print(f"第{page_num}保存图片完成")
            page_num += 1
    return img_outs
